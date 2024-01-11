import cv2
from datetime import datetime, date
from openpyxl import *
import uc2rest as uc2
import numpy as np
from openpyxl.styles import Font

# Constants for Rows and Framing
max_rows = 3
max_colums = 3
row = 0
column = 0
x_frame = 1670
y_frame = 980
left_far = 0

# Create Excel Workbook and Active Sheet
Excel_file = Workbook()
ws = Excel_file.active

# Define new frame dimensions
y1 = 100
y2 = 1080 - y1
x1 = 250
x2 = 1920 - x1

# Setup for Contrast Limited Adaptive Histogram Equalization (CLAHE)
clahe = cv2.createCLAHE(clipLimit=5.0, tileGridSize=(8, 8))
connectivity = 8

# Initial Movement Values
X_movement = 0
Y_movement = 0

# Serial port for communication with the board
serialport = "COM5"

# Check if ESP32 object exists, create if not
if 'ESP32' not in locals():
    ESP32 = uc2.UC2Client(serialport=serialport)

# Retrieve and print the current state of the ESP32 board
_state = ESP32.state.get_state()
print("Current State:", _state)

# Motor setup (executed only if the condition is True)
if 0:
    ESP32.motor.set_motor(stepperid=1, position=0, stepPin=26, dirPin=16, enablePin=12, maxPos=None, minPos=None,
                          acceleration=None, isEnable=1)
    ESP32.motor.set_motor(stepperid=2, position=0, stepPin=25, dirPin=27, enablePin=12, maxPos=None, minPos=None,
                          acceleration=None, isEnable=1)
    ESP32.motor.set_motor(stepperid=3, position=0, stepPin=17, dirPin=14, enablePin=12, maxPos=None, minPos=None,
                          acceleration=None, isEnable=1)
    ESP32.motor.set_motor(stepperid=0, position=0, stepPin=19, dirPin=18, enablePin=12, maxPos=None, minPos=None,
                          acceleration=None, isEnable=1)


def Move_X(steps):
    """
    Move along the X-axis.

    Parameters:
    - steps: Number of steps to move (negative for reverse direction).
    """
    ESP32.motor.move_xyz(
        steps=(-steps, 2 * steps, -steps),
        speed=(10000, 10000, 10000),
        acceleration=None,
        is_blocking=False,
        is_absolute=False,
        is_enabled=True
    )


def Move_Y(steps):
    """
    Move along the Y-axis.

    Parameters:
    - steps: Number of steps to move (negative for reverse direction).
    """
    ESP32.motor.move_xyz(
        steps=(steps, 0, -steps),
        speed=(10000, 10000, 10000),
        acceleration=None,
        is_blocking=False,
        is_absolute=False,
        is_enabled=True
    )


def Move_Z(steps):
    """
    Move along the Z-axis.

    Parameters:
    - steps: Number of steps to move (negative for reverse direction).
    """
    ESP32.motor.move_xyz(
        steps=(steps, steps, steps),
        speed=(10000, 10000, 10000),
        acceleration=None,
        is_blocking=False,
        is_absolute=False,
        is_enabled=True
    )

def file_generation(max_columns, row_nr, column_nr, summation):
    """
    Determine the next position and movement values in a grid.

    Parameters:
    - max_columns: Maximum number of columns in the grid.
    - row_nr: Current row number.
    - column_nr: Current column number.
    - summation: Cumulative summation value.

    Returns:
    - Tuple containing (position, updated row number, updated column number, move value, super move flag, updated summation).
    """
    # Check for a special condition that triggers super move (aka going down a row)
    if column_nr + 1 == max_columns - 1 and ((row_nr % 2) == 0) or column_nr - 1 == 0 and ((row_nr % 2) != 0):
        super_move = 1
    else:
        super_move = 0

    # Update position based on current column and row conditions
    if (column_nr == max_columns - 1) and ((row_nr % 2) == 0):
        column_nr += 1
        row_nr += 1

    if (column_nr == 0) and ((row_nr % 2) != 0):
        column_nr -= 1
        row_nr += 1

    # Move one column to the right for even rows
    if ((row_nr % 2) == 0):
        column_nr += 1
        move_val = 1

    # Move one column to the left for odd rows
    if ((row_nr % 2) != 0):
        column_nr -= 1
        move_val = 2
        if column_nr != max_columns - 1:
            summation += 2

    # Reset move value if it's a downward move
    if super_move == 1:
        move_val = 0

    return (str(row_nr) + ";" + str(column_nr), row_nr, column_nr, move_val, super_move, summation)


def how_late_is_it():
    """
    Get the current date and time.

    Returns:
    - Tuple containing current date in dd_mm format and current time in HH;MM;SS format.
    """
    today = date.today()
    d1 = today.strftime("%d_%m")
    now = datetime.now()
    current_time = now.strftime("%H;%M;%S")
    return d1, current_time


def Edge_Detection_Closing(img, file_path_segment):
    """
    Perform edge detection and closing operation on an image.

    Parameters:
    - img: Input image.
    - file_path_segment: File path to save the processed image.
    """
    # Calculate histogram and find thresholds for binary image
    dst = cv2.calcHist([img], [0], None, [256], [0, 256])

    value_list_dark = []
    value_list_light = []
    for i in range(90):
        value_list_dark.append(int(dst[i][0]))

    for i in range(len(dst)):
        value_list_light.append(int(dst[i][0]))

    final_val_dark = value_list_dark.index(max(value_list_dark))
    final_val_light = value_list_light.index(max(value_list_light))

    # Threshold the image
    ret, binary_img = cv2.threshold(img, final_val_dark, final_val_light, cv2.THRESH_BINARY)

    # Apply Sobel operator for edge detection
    sobelxy = cv2.Sobel(src=binary_img, ddepth=cv2.CV_64F, dx=1, dy=1, ksize=13)

    ret, sobelxy = cv2.threshold(sobelxy, 0, 256, cv2.THRESH_BINARY)

    # Create kernel for morphological operations
    kernel = np.ones((8, 8), np.uint8)
    gradient = cv2.morphologyEx(sobelxy, cv2.MORPH_GRADIENT, kernel)

    kernel_begin = np.ones((5, 5), np.uint8)

    # Apply closing operation
    closing_1 = cv2.morphologyEx(gradient, cv2.MORPH_CLOSE, kernel_begin)
    cv2.imwrite(file_path_segment, closing_1)


def Excel_generator(ws):
    """
    Generate a new Excel file with specific formatting.

    Parameters:
    - ws: Excel sheet object.

    Returns:
    - Updated Excel file object.
    """
    Exfile = ws
    # Set title based on current date and time
    Exfile.title = "Run" + how_late_is_it()[0] + "_Time_" + how_late_is_it()[1]
    ws.cell(1, 1).value = "X"
    ws.cell(1, 1).font = Font(bold=True)
    ws.cell(1, 2).value = "Y"
    ws.cell(1, 2).font = Font(bold=True)

    return Exfile

def Excel_writer(worksheet, imx, imy, x_frame, y_frame, cords, summation):
    """
    Write coordinates to an Excel worksheet.

    Parameters:
    - worksheet: Excel worksheet object.
    - imx: Image x-coordinate multiplier.
    - imy: Image y-coordinate multiplier.
    - x_frame: Frame x-coordinate.
    - y_frame: Frame y-coordinate.
    - cords: List of coordinates to be written.
    - summation: Cumulative summation value.

    Returns:
    - Updated summation value.
    """
    for i in range(len(cords)):
        # Calculate adjusted coordinates based on image and frame parameters
        x_coordinate = x_frame * imx + cords[i][0]
        y_coordinate = y_frame * imy + cords[i][1]

        # Write coordinates to the Excel worksheet
        cellx = worksheet.cell(column=1, row=(imx + imy + i + summation))
        celly = worksheet.cell(column=2, row=(imx + imy + i + summation))
        cellx.value = x_coordinate
        celly.value = y_coordinate

    # Update summation for the next iteration
    summation += len(cords) - 1
    return summation


worksheet = Excel_generator(ws)
summation = 2

# File paths for saving images
Original_filename = str(row) + ";" + str(column)
File_Pathing = "C:/Users/20202619/OneDrive - TU Eindhoven/Vakken/CBL Microscopy/Python_Code_Microscope_Redux/Full_Image_Run/"
File_Pathing_Excel = "C:/Users/20202619/OneDrive - TU Eindhoven/Vakken/CBL Microscopy/Python_Code_Microscope_Redux/"
File_Pathing_Detected = "C:/Users/20202619/OneDrive - TU Eindhoven/Vakken/CBL Microscopy/Python_Code_Microscope_Redux/Full_Image_Run_Dots/"
File_Pathing_Segment = "C:/Users/20202619/OneDrive - TU Eindhoven/Vakken/CBL Microscopy/Python_Code_Microscope_Redux/Full_Image_Run_Segment/"
File_Pathing_Dots = "C:/Users/20202619/OneDrive - TU Eindhoven/Vakken/CBL Microscopy/Python_Code_Microscope_Redux/Full_Image_Run_Dots_Add/"
file_path = File_Pathing + Original_filename + ".png"
file_path_detected = File_Pathing_Detected + Original_filename + ".png"
file_path_segment = File_Pathing_Segment + Original_filename + ".png"


# Move into the top-left position
Move_X(-36000)
Move_Y(79050)
# Record movements
X_movement -= 36000
Y_movement += 79050

while True:
    # Capture video from the default camera (index 0)
    vid = cv2.VideoCapture(0)

    # Generate image name based on row and column
    Image_name = Original_filename
    file_path = File_Pathing + Image_name + ".png"  # Filepath where image is written
    file_path_detected = File_Pathing_Detected + Image_name + "dots" + ".png"
    file_path_segment = File_Pathing_Segment + Image_name + "Segment" + ".png"

    # Read and resize the video frame
    ret, frame = vid.read()
    frame = cv2.resize(frame, (1920, 1080))
    grayFrame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    FrameNewRes = grayFrame[y1:y2, x1:x2]
    cv2.imwrite(file_path, FrameNewRes)

    # Move the microscope to the right
    Move_X(24000)
    X_movement += 24000
    print("Move Right")

    # Apply contrast enhancement and edge detection
    Frame_Con = clahe.apply(FrameNewRes)
    Edge_Detection_Closing(Frame_Con, file_path_segment)
    Segmented_Img = cv2.imread(file_path_segment, cv2.IMREAD_GRAYSCALE)

    # Extract centroids of detected objects
    white_list = np.zeros(Segmented_Img.shape)
    output = cv2.connectedComponentsWithStats(Segmented_Img, connectivity, cv2.CV_32S)
    centroids = np.array(output[3])

    centroids_rounded = (np.rint(centroids)).astype(int)
    for coords in centroids_rounded:
        x = coords[0]
        y = coords[1]
        white_list[y, x] = 255

    white_list = white_list.astype(np.uint8)

    # Apply dilation to the detected dots
    kernel_dilate = np.ones((4, 4), np.uint8)
    dilated_dots = cv2.dilate(white_list, kernel_dilate, iterations=3)
    cv2.imwrite(file_path_detected, dilated_dots)

    # Write coordinates to Excel file
    summation = Excel_writer(worksheet, 0, 0, x_frame, y_frame, centroids_rounded, summation)
    Excel_file.save(File_Pathing_Excel + "Coordinate_Sheet.xlsx")

    # Release video capture object and close all windows
    vid.release()
    cv2.destroyAllWindows()

    # Loop for grid movement
    while (row != (max_rows - 1)) or (column != (max_colums - 1)):
        vid = cv2.VideoCapture(0)

        # Generate new image name and file paths
        Image_name, row, column, move_val, super_move, summation = file_generation(max_colums, row, column, summation)
        file_path = File_Pathing + Image_name + ".png"
        file_path_detected = File_Pathing_Detected + Image_name + "dots" + ".png"
        file_path_segment = File_Pathing_Segment + Image_name + "_Segment" + ".png"

        # Read and resize the video frame
        ret, frame = vid.read()
        frame = cv2.resize(frame, (1920, 1080))
        grayFrame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        FrameNewRes = grayFrame[y1:y2, x1:x2]

        cv2.imwrite(file_path, FrameNewRes)

        # Move the microscope based on the specified conditions
        if super_move == 1:
            Move_Y(-52700)
            Y_movement -= 52700
            print("Move Down")
        elif move_val == 1:
            Move_X(24000)
            X_movement += 24000
            print("Move Right")
        elif move_val == 2:
            Move_X(-24000)
            X_movement -= 24000
            print("Move Left")

        # Apply contrast enhancement and edge detection
        Frame_Con = clahe.apply(FrameNewRes)
        Edge_Detection_Closing(Frame_Con, file_path_segment)
        Segmented_Img = cv2.imread(file_path_segment, cv2.IMREAD_GRAYSCALE)

        # Extract centroids of detected objects
        white_list = np.zeros(Segmented_Img.shape)
        output = cv2.connectedComponentsWithStats(Segmented_Img, connectivity, cv2.CV_32S)
        centroids = np.array(output[3])

        centroids_rounded = (np.rint(centroids)).astype(int)
        for coords in centroids_rounded:
            x = coords[0]
            y = coords[1]
            white_list[y, x] = 255

        white_list = white_list.astype(np.uint8)

        # Apply dilation to the detected dots
        kernel_dilate = np.ones((4, 4), np.uint8)
        dilated_dots = cv2.dilate(white_list, kernel_dilate, iterations=3)
        cv2.imwrite(file_path_detected, dilated_dots)

        # Write coordinates to Excel file
        summation = Excel_writer(worksheet, column, row, x_frame, y_frame, centroids_rounded, summation)
        Excel_file.save(File_Pathing_Excel + "Coordinate_Sheet.xlsx")

        # Release video capture object and close all windows
        vid.release()
        cv2.destroyAllWindows()

    # Break out of the main loop
    break

# Move the microscope back to the original position
Move_X(-1 * X_movement)
Move_Y(-1 * Y_movement)