import cv2
from datetime import datetime, date
from openpyxl import *
from openpyxl.styles import Font

max_rows = 5  #number of rows
max_colums = 5 #number of columns
row = 0
column = 0
x_frame = 1670
y_frame = 980
example_dataset = [[100,300],[200,500],[1000,1000]]

#make excel files
Excel_file = Workbook()
ws = Excel_file.active


def file_generation(max_columns,row_nr,column_nr):
        
    if (column_nr == max_columns-1) and ((row_nr % 2)==0): #check if row has already been shifted down and change direction to right
        column_nr += 1
        row_nr += 1
    
    if (column_nr == 0) and ((row_nr % 2)!=0): #check if row has already been shifted down and change direction to right
        column_nr -= 1
        row_nr += 1

    if ((row_nr % 2)==0): #one col to the right
        column_nr += 1 
            
    if ((row_nr % 2)!=0): #one col to the left
        column_nr -= 1
        
    return((str(row_nr) + ";" + str(column_nr)),row_nr,column_nr)
def how_late_is_it():
    today = date.today()
    d1 = today.strftime("%d_%m")
    now = datetime.now()
    current_time = now.strftime("%H;%M;%S")
    return(d1,current_time)
def Excel_generator(ws):
    Exfile = ws
    Exfile.title = "Run" + how_late_is_it()[0] + "_Time_" + how_late_is_it()[1]
    ws.cell(1,1).value = "X"
    ws.cell(1,1).font = Font(bold=True)
    ws.cell(1,2).value = "Y"
    ws.cell(1,2).font = Font(bold=True)

    return(Exfile)
def Excel_writer(worksheet,imx,imy,x_frame,y_frame,cords,summation):
    for i in range(len(cords)):
        x_coordinate = x_frame * imx + cords[i][0]
        y_coordinate = y_frame * imy + cords[i][1]
        cellx = worksheet.cell(column=1,row=(imx+imy+i+summation))
        celly = worksheet.cell(column=2,row=(imx+imy+i+summation))
        cellx.value = x_coordinate
        celly.value = y_coordinate
    summation += len(cords)-1 
    return(summation)
    
Original_filename = str(row) + ";" + str(column)
File_Pathing = "C:/Users/20202619/OneDrive - TU Eindhoven/Vakken/CBL Microscopy/Python_Code_Microscope_Redux/Full_Image_Run/"
file_path = File_Pathing + Original_filename + ".png" #Filepath where image is written

while(True):

    vid = cv2.VideoCapture(0)
    Image_name = Original_filename
    file_path = File_Pathing + Image_name + ".png" #Filepath where image is written
    ret, frame = vid.read()
    frame = cv2.resize(frame, (1920, 1080))
    grayFrame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

    if cv2.waitKey(1):     
        cv2.imwrite(file_path, grayFrame)
        vid.release() # After the loop release the cap object
        cv2.destroyAllWindows() # Destroy all the windows

    while (row != (max_rows-1)) or (column != (max_colums-1)):
        vid = cv2.VideoCapture(0)
        Image_name,row,column = file_generation(max_colums,row,column)
        file_path = File_Pathing + Image_name + ".png" #Filepath where image is written
        ret, frame = vid.read()
        frame = cv2.resize(frame, (1920, 1080))
        grayFrame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

        if cv2.waitKey(1):     
            cv2.imwrite(file_path, grayFrame)
            vid.release() # After the loop release the cap object
            cv2.destroyAllWindows() # Destroy all the windows
    break 

#Analysis part of the code
summation = 2
worksheet = Excel_generator(ws)
#for loop can go over each image
for ex_row in range(max_rows):
    for ex_col in range(max_colums):
        summation = Excel_writer(worksheet,ex_col,ex_row,x_frame,y_frame,example_dataset,summation)
    summation += (max_colums-1)  
Excel_file.save((File_Pathing+"Hello1.xlsx"))











    





