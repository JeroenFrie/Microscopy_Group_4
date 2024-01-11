import cv2
from datetime import datetime, date
from openpyxl import *
import uc2rest as uc2
import numpy as np
from openpyxl.styles import Font

#Rows and Framing
max_rows = 3
max_colums = 3
row = 0
column = 0
x_frame = 1670
y_frame = 980
left_far = 0

#make excel files
Excel_file = Workbook()
ws = Excel_file.active


#New frame
y1 = 100
y2 = 1080-y1
x1 = 250
x2 = 1920-x1

#Contrast Setup
clahe = cv2.createCLAHE(clipLimit=5.0, tileGridSize=(8, 8))
connectivity = 8

#Movement Values
X_movement = 0
Y_movement = 0

#Make a connection with the board
serialport = "COM5"

if 'ESP32' not in locals():
    ESP32 = uc2.UC2Client(serialport=serialport)
_state = ESP32.state.get_state()
print(_state)

if 0:
    ESP32.motor.set_motor(stepperid = 1, position = 0, stepPin = 26, dirPin=16, enablePin=12, maxPos=None, minPos=None, acceleration=None, isEnable=1)
    ESP32.motor.set_motor(stepperid = 2, position = 0, stepPin = 25, dirPin=27, enablePin=12, maxPos=None, minPos=None, acceleration=None, isEnable=1)
    ESP32.motor.set_motor(stepperid = 3, position = 0, stepPin = 17, dirPin=14, enablePin=12, maxPos=None, minPos=None, acceleration=None, isEnable=1)
    ESP32.motor.set_motor(stepperid = 0, position = 0, stepPin = 19, dirPin=18, enablePin=12, maxPos=None, minPos=None, acceleration=None, isEnable=1)


def Move_X(steps):
    ESP32.motor.move_xyz(
        steps=(-steps, 2 * steps, -steps),
        speed=(10000, 10000, 10000),
        acceleration=None,
        is_blocking=False,
        is_absolute=False,
        is_enabled=True
    )


def Move_Y(steps):
    ESP32.motor.move_xyz(
        steps=(steps, 0, -steps),
        speed=(10000, 10000, 10000),
        acceleration=None,
        is_blocking=False,
        is_absolute=False,
        is_enabled=True
    )


def Move_Z(steps):
    ESP32.motor.move_xyz(
        steps=(steps, steps, steps),
        speed=(10000, 10000, 10000),
        acceleration=None,
        is_blocking=False,
        is_absolute=False,
        is_enabled=True
    )

def file_generation(max_columns,row_nr,column_nr,summation):
    if column_nr+1 == max_columns-1 and ((row_nr % 2)==0) or column_nr-1 == 0 and ((row_nr % 2)!=0):
        super_move = 1

    else:
        super_move = 0

    if (column_nr == max_columns-1) and ((row_nr % 2)==0): #check if row has already been shifted down and change direction to right
        column_nr += 1
        row_nr += 1

    if (column_nr == 0) and ((row_nr % 2)!=0): #check if row has already been shifted down and change direction to right
        column_nr -= 1
        row_nr += 1

    if ((row_nr % 2)==0): #one col to the right
        column_nr += 1 
        move_val = 1

    if ((row_nr % 2)!=0): #one col to the left
        column_nr -= 1
        move_val = 2
        if column_nr != max_columns-1:
            summation += 2

    if super_move == 1:
        move_val = 0


    return((str(row_nr) + ";" + str(column_nr)),row_nr,column_nr,move_val,super_move,summation)

def how_late_is_it():
    today = date.today()
    d1 = today.strftime("%d_%m")
    now = datetime.now()
    current_time = now.strftime("%H;%M;%S")
    return(d1,current_time)

def Edge_Detection_Closing(img,file_path_segment):
    dst = cv2.calcHist([img], [0], None, [256], [0, 256])
    max_val = 0
    list_dst = list(dst)

    value_list_dark = []
    value_list_light = []
    for i in range(90):
        value_list_dark.append(int(dst[i][0]))

    for i in range(len(dst)):
        value_list_light.append(int(dst[i][0]))

    final_val_dark = value_list_dark.index(max(value_list_dark))
    final_val_light = value_list_light.index(max(value_list_light))

    ret, binary_img = cv2.threshold(img, final_val_dark, final_val_light, cv2.THRESH_BINARY)

    sobelxy = cv2.Sobel(src=binary_img, ddepth=cv2.CV_64F, dx=1, dy=1, ksize=13)

    ret, sobelxy = cv2.threshold(sobelxy, 0, 256, cv2.THRESH_BINARY)

    kernel = np.ones((8, 8), np.uint8)
    gradient = cv2.morphologyEx(sobelxy, cv2.MORPH_GRADIENT, kernel)

    kernel_begin = np.ones((5, 5), np.uint8)

    closing_1 = cv2.morphologyEx(gradient, cv2.MORPH_CLOSE, kernel_begin)
    cv2.imwrite(file_path_segment, closing_1)

def Excel_generator(ws):
    Exfile = ws
    Exfile.title = "Run" + how_late_is_it()[0] + "_Time_" + how_late_is_it()[1]
    ws.cell(1,1).value = "X"
    ws.cell(1,1).font = Font(bold=True)
    ws.cell(1,2).value = "Y"
    ws.cell(1,2).font = Font(bold=True)

    return(Exfile)

def Excel_writer(worksheet,imx,imy,x_frame,y_frame,cords,summation):
    for i in range(len(cords)):
        x_coordinate = x_frame * imx + cords[i][0]
        y_coordinate = y_frame * imy + cords[i][1]
        cellx = worksheet.cell(column=1,row=(imx+imy+i+summation))
        celly = worksheet.cell(column=2,row=(imx+imy+i+summation))
        cellx.value = x_coordinate
        celly.value = y_coordinate
    summation += len(cords)-1
    return(summation)

worksheet = Excel_generator(ws)
summation = 2

Original_filename = str(row) + ";" + str(column)
File_Pathing = "C:/Users/20202619/OneDrive - TU Eindhoven/Vakken/CBL Microscopy/Python_Code_Microscope_Redux/Full_Image_Run/"
File_Pathing_Excel = "C:/Users/20202619/OneDrive - TU Eindhoven/Vakken/CBL Microscopy/Python_Code_Microscope_Redux/"
File_Pathing_Detected = "C:/Users/20202619/OneDrive - TU Eindhoven/Vakken/CBL Microscopy/Python_Code_Microscope_Redux/Full_Image_Run_Dots/"
File_Pathing_Segment = "C:/Users/20202619/OneDrive - TU Eindhoven/Vakken/CBL Microscopy/Python_Code_Microscope_Redux/Full_Image_Run_Segment/"
File_Pathing_Dots = "C:/Users/20202619/OneDrive - TU Eindhoven/Vakken/CBL Microscopy/Python_Code_Microscope_Redux/Full_Image_Run_Dots_Add/"
file_path = File_Pathing + Original_filename + ".png" #Filepath where image is written
file_path_detected = File_Pathing_Detected + Original_filename + ".png"
file_path_segment = File_Pathing_Segment + Original_filename + ".png"
file_path_dots = File_Pathing_Dots + Original_filename + ".png"




#Move into top left
Move_X(-36000)
Move_Y(79050)
X_movement -= 36000
Y_movement += 79050
while(True):

    vid = cv2.VideoCapture(0)
    Image_name = Original_filename
    file_path = File_Pathing + Image_name + ".png" #Filepath where image is written
    file_path_detected = File_Pathing_Detected + Image_name + "dots" + ".png"
    file_path_segment = File_Pathing_Segment + Image_name + "Segment" + ".png"
    file_path_dots = File_Pathing_Dots + Image_name + ".png"
    ret, frame = vid.read()
    frame = cv2.resize(frame, (1920, 1080))
    grayFrame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    FrameNewRes = grayFrame[y1:y2, x1:x2]
    cv2.imwrite(file_path, FrameNewRes)

    Move_X(24000)
    X_movement += 24000
    print("Move Right")

    Frame_Con = clahe.apply(FrameNewRes)
    Edge_Detection_Closing(Frame_Con,file_path_segment)
    Segmented_Img = cv2.imread(file_path_segment, cv2.IMREAD_GRAYSCALE)

    white_list = np.zeros(Segmented_Img.shape)
    output = cv2.connectedComponentsWithStats(Segmented_Img, connectivity, cv2.CV_32S)
    centroids = np.array(output[3])

    centroids_rounded = (np.rint(centroids)).astype(int)
    for coords in centroids_rounded:
        x = coords[0]
        y = coords[1]
        white_list[y, x] = 255
        #Dotted_Image = FrameNewRes[y, x] = 255

    #cv2.imwrite(file_path_dots, Dotted_Image)
    white_list = white_list.astype(np.uint8)

    kernel_dilate = np.ones((4, 4), np.uint8)
    dilated_dots = cv2.dilate(white_list, kernel_dilate, iterations=3)
    cv2.imwrite(file_path_detected, dilated_dots)
    example_dataset = [[0, 0], [500, 500], [1000, 1000]]
    summation = Excel_writer(worksheet, 0, 0, x_frame, y_frame, centroids_rounded, summation)
    Excel_file.save((File_Pathing_Excel + "Hello.xlsx"))

    vid.release()  # After the loop release the cap object
    cv2.destroyAllWindows()  # Destroy all the windows


    while (row != (max_rows-1)) or (column != (max_colums-1)):
        vid = cv2.VideoCapture(0)
        Image_name,row,column,move_val,super_move,summation = file_generation(max_colums,row,column,summation)
        file_path = File_Pathing + Image_name + ".png"
        file_path_detected = File_Pathing_Detected + Image_name + "dots" + ".png"#Filepath where image is written
        file_path_segment = File_Pathing_Segment + Image_name + "_Segment" + ".png"
        file_path_dots = File_Pathing_Dots + Image_name + ".png"
        ret, frame = vid.read()
        frame = cv2.resize(frame, (1920, 1080))
        grayFrame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        FrameNewRes = grayFrame[y1:y2, x1:x2]

        cv2.imwrite(file_path, FrameNewRes)

        if super_move == 1:
            Move_Y(-52700)
            Y_movement -= 52700
            print("Move Down")
        elif move_val == 1:
            Move_X(24000)
            X_movement += 24000
            print("Move Right")
        elif move_val == 2:
            Move_X(-24000)
            X_movement -= 24000
            print("Move Left")


        Frame_Con = clahe.apply(FrameNewRes)
        Edge_Detection_Closing(Frame_Con,file_path_segment)
        Segmented_Img = cv2.imread(file_path_segment, cv2.IMREAD_GRAYSCALE)
        white_list = np.zeros(Segmented_Img.shape)
        output = cv2.connectedComponentsWithStats(Segmented_Img, connectivity, cv2.CV_32S)
        centroids = np.array(output[3])

        centroids_rounded = (np.rint(centroids)).astype(int)
        for coords in centroids_rounded:
            x = coords[0]
            y = coords[1]
            white_list[y, x] = 255
            #Dotted_Image = FrameNewRes[y, x] = 255

        #cv2.imwrite(file_path_dots, Dotted_Image)
        white_list = white_list.astype(np.uint8)

        kernel_dilate = np.ones((4, 4), np.uint8)
        dilated_dots = cv2.dilate(white_list, kernel_dilate, iterations=3)
        cv2.imwrite(file_path_detected, dilated_dots)

        example_dataset = [[0, 0], [500, 500], [1000, 1000]]
        summation = Excel_writer(worksheet, column, row, x_frame, y_frame, centroids_rounded, summation)
        Excel_file.save((File_Pathing_Excel + "Hello.xlsx"))
        vid.release()  # After the loop release the cap object
        cv2.destroyAllWindows()  # Destroy all the windows

    break


Move_X(-1*X_movement)
Move_Y(-1*Y_movement)